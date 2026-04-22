"""Import-template generators.

Emit a CSV and XLSX file with the full header row the importer accepts,
plus 3 demo rows that show realistic values. Operators download these,
fill them in with their own data, and re-upload via /requirements/import/.
"""
import csv
import io

from tcms_requirements.exports.csv_export import COLUMNS


SAMPLE_ROWS = [
    {
        "identifier": "SYS-REQ-001",
        "title": "The system shall start within 3 seconds of power-on.",
        "description": "Cold boot time from power-on to interactive shell.",
        "rationale": "Improves first-use impression and meets customer RFP §3.2.",
        "category": "Performance",
        "source": "Default customer document",
        "source_section": "§3.2",
        "level": "system",
        "product": "",
        "project": "",
        "feature": "",
        "parent_requirement": "",
        "status": "approved",
        "priority": "high",
        "verification_method": "test",
        "verification_exemption_reason": "",
        "asil": "",
        "sil": "",
        "iec62304_class": "",
        "dal": "",
        "doc_id": "SRS-2026-A",
        "doc_revision": "A",
        "effective_date": "2026-05-01",
        "superseded_by": "",
        "change_reason": "",
        "jira_issue_key": "",
        "external_refs": "",
        "linked_cases": "",
    },
    {
        "identifier": "SW-REQ-042",
        "title": "Login screen shall lock after 5 failed attempts.",
        "description": "Progressive back-off after repeated failed logins.",
        "rationale": "Mitigates brute-force per OWASP ASVS 2.2.",
        "category": "Security",
        "source": "Default technical specification",
        "source_section": "§5.1",
        "level": "software",
        "product": "",
        "project": "",
        "feature": "",
        "parent_requirement": "SYS-REQ-001",
        "status": "draft",
        "priority": "critical",
        "verification_method": "test",
        "verification_exemption_reason": "",
        "asil": "",
        "sil": "",
        "iec62304_class": "",
        "dal": "",
        "doc_id": "",
        "doc_revision": "",
        "effective_date": "",
        "superseded_by": "",
        "change_reason": "",
        "jira_issue_key": "",
        "external_refs": "",
        "linked_cases": "",
    },
    {
        "identifier": "CUST-NFR-017",
        "title": "The product shall display the corporate logo on every page.",
        "description": "Brand identity requirement (non-functional).",
        "rationale": "Customer brand-guidelines compliance.",
        "category": "UI/UX",
        "source": "Default customer document",
        "source_section": "§2.1",
        "level": "stakeholder",
        "product": "",
        "project": "",
        "feature": "",
        "parent_requirement": "",
        "status": "verified",
        "priority": "medium",
        "verification_method": "inspection",
        "verification_exemption_reason": "",
        "asil": "",
        "sil": "",
        "iec62304_class": "",
        "dal": "",
        "doc_id": "",
        "doc_revision": "",
        "effective_date": "",
        "superseded_by": "",
        "change_reason": "",
        "jira_issue_key": "",
        "external_refs": "",
        "linked_cases": "",
    },
]


def write_csv_template(buf) -> None:
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, quoting=csv.QUOTE_MINIMAL)
    writer.writeheader()
    for row in SAMPLE_ROWS:
        writer.writerow({col: row.get(col, "") for col in COLUMNS})


def build_xlsx_template() -> bytes:
    """Return an XLSX workbook as bytes.

    Uses openpyxl's in-memory write support. Freezes the header row so
    operators scrolling long sheets still see column names.
    """
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="39A5DC", end_color="39A5DC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, sample in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample.get(name, ""))

    ws.freeze_panes = "A2"
    # Approximate sensible widths for the most common columns.
    widths = {
        "identifier": 18, "title": 42, "description": 40, "rationale": 30,
        "category": 14, "source": 24, "level": 12, "product": 14,
        "project": 14, "feature": 14, "status": 14, "priority": 10,
        "verification_method": 18, "doc_id": 16, "jira_issue_key": 14,
    }
    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(name, 18)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
