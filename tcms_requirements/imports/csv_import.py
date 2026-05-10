"""Bulk import with dry-run preview.

Accepts the same column set `exports/csv_export.py::COLUMNS` produces, so
export → edit → import round-trips cleanly. Dry-run validates everything
in a transaction that's rolled back, reporting counts + errors without
touching the DB. A non-dry-run commits.

Two input formats:
 - CSV (any UTF-8 dialect `csv.Sniffer` can parse)
 - XLSX (first worksheet, first row as header)

FKs resolved by human-friendly names:
    category → RequirementCategory.name
    source → RequirementSource.name (first match)
    level → RequirementLevel.code
    product → Product.name
    project → Project.name
    feature → Feature.name
    parent_requirement / superseded_by → Requirement.identifier

Test-case links via the ``linked_cases`` column:
    Comma-separated TestCase IDs (e.g. ``"42,57,103"``) — bare integers
    or ``TC-42`` prefixed (the JIRA-style export). On update, existing
    links are *replaced* with the column's contents so the round-trip is
    deterministic. ``link_type`` defaults to ``verifies`` and ``suspect``
    to ``False`` because those signals don't survive the flat export
    (use the JSON exporter / JSON importer when full link fidelity is
    required).

Missing FK targets → error for that row; row is skipped but the import
continues so the user sees every error in one pass.
"""
import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field

from django.db import transaction

from tcms_requirements.models import (
    Feature,
    Project,
    Requirement,
    RequirementCategory,
    RequirementLevel,
    RequirementSource,
    RequirementTestCaseLink,
)

logger = logging.getLogger("tcms_requirements")


REQUIRED_COLUMNS = {"identifier", "title"}

# "TC-42" or bare "42" — both accepted, comma-separated.
_CASE_ID_PATTERN = re.compile(r"^\s*(?:TC-)?(\d+)\s*$", re.IGNORECASE)


@dataclass
class RowError:
    row_num: int
    identifier: str
    message: str


@dataclass
class ImportResult:
    rows_total: int = 0
    rows_ok: int = 0
    rows_created: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    errors: list = field(default_factory=list)


# ── dispatch ─────────────────────────────────────────────────────────


def import_bytes(data: bytes, filename: str, *, dry_run: bool = True, user=None) -> ImportResult:
    """Auto-detect input format and dispatch.

    Recognised by filename extension:
      * ``.json`` → JSON importer (full link-fidelity round-trip)
      * ``.xlsx`` / ``.xlsm`` → XLSX worksheet
      * everything else → CSV (UTF-8, any dialect csv.Sniffer can parse)

    For CSV, JIRA-style column headers (``External Issue ID``,
    ``Summary``, ``Linked Test Cases``, etc.) are auto-translated to the
    canonical schema so a JIRA-CSV export round-trips without the user
    having to rename headers by hand.
    """
    lower = (filename or "").lower()
    if lower.endswith(".json"):
        from tcms_requirements.imports.json_import import import_json_bytes  # noqa: WPS433
        return import_json_bytes(data, dry_run=dry_run, user=user)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        rows, fieldnames = _read_xlsx(data)
    else:
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            return _error_result("File must be UTF-8 encoded CSV, JSON, or .xlsx.")
        reader = csv.DictReader(io.StringIO(text))
        fieldnames = reader.fieldnames or []
        rows = list(reader)
    rows, fieldnames = _normalise_jira_columns(rows, fieldnames)
    return _import_rows(rows, fieldnames, dry_run=dry_run, user=user)


def import_csv(raw_text: str, *, dry_run: bool = True, user=None) -> ImportResult:
    """Legacy entry point — callers that already have a CSV string."""
    reader = csv.DictReader(io.StringIO(raw_text))
    fieldnames = reader.fieldnames or []
    rows = list(reader)
    rows, fieldnames = _normalise_jira_columns(rows, fieldnames)
    return _import_rows(rows, fieldnames, dry_run=dry_run, user=user)


# ── JIRA CSV column aliasing ─────────────────────────────────────────


# JIRA "External System Import" column names → our canonical names.
# We keep the JIRA columns we know how to round-trip; the rest pass
# through unchanged so unknown headers become inert (ignored).
_JIRA_TO_CANONICAL = {
    "External Issue ID": "identifier",
    "Issue Key": "jira_issue_key",
    "Summary": "title",
    "Description": "description",
    "Priority": "priority",
    "Status": "status",
    "Component/s": "category",
    "Linked Test Cases": "linked_cases",
    "Parent Requirement": "parent_requirement",
    "Source Document": "doc_id",
    "Document File Name": "document_file_name",
    "Document Title": "document_title",
    "ASIL": "asil",
    "DAL": "dal",
    "IEC 62304 Class": "iec62304_class",
    "Requirement Level": "level",
}

# Reverse maps for the value translations the JIRA exporter applies.
_JIRA_PRIORITY_REVERSE = {
    "Highest": "critical", "High": "high", "Medium": "medium", "Low": "low",
    "Lowest": "low",
}
_JIRA_STATUS_REVERSE = {
    "To Do": "draft", "In Progress": "in_review", "Done": "approved",
    "Closed": "deprecated",
}


def _looks_like_jira_csv(fieldnames) -> bool:
    """Heuristic: JIRA-CSV always carries both Summary and either
    External Issue ID or Issue Key, none of which appear in our generic
    CSV. The check is order-insensitive and case-sensitive (JIRA's
    headers are stable)."""
    fields = set(fieldnames or [])
    return "Summary" in fields and (
        "External Issue ID" in fields or "Issue Key" in fields
    )


def _normalise_jira_columns(rows, fieldnames):
    """If the input looks like a JIRA CSV, rename headers to canonical
    names and reverse-translate the priority/status enums.

    Returns ``(rows, fieldnames)`` — both with canonical names.
    """
    if not _looks_like_jira_csv(fieldnames):
        return rows, fieldnames

    new_fieldnames = [
        _JIRA_TO_CANONICAL.get(name, name)
        for name in (fieldnames or [])
    ]
    new_rows = []
    for row in rows:
        translated = {}
        for old_key, value in row.items():
            new_key = _JIRA_TO_CANONICAL.get(old_key, old_key)
            translated[new_key] = value
        if translated.get("priority") in _JIRA_PRIORITY_REVERSE:
            translated["priority"] = _JIRA_PRIORITY_REVERSE[translated["priority"]]
        if translated.get("status") in _JIRA_STATUS_REVERSE:
            translated["status"] = _JIRA_STATUS_REVERSE[translated["status"]]
        new_rows.append(translated)
    return new_rows, new_fieldnames


# ── XLSX reader ──────────────────────────────────────────────────────


def _read_xlsx(data: bytes):
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration:
        return [], []
    fieldnames = [str(h).strip() if h is not None else "" for h in header]
    rows = []
    for values in iterator:
        if all(v is None or str(v).strip() == "" for v in values):
            continue  # skip blank rows
        row = {}
        for idx, col in enumerate(fieldnames):
            if not col:
                continue
            if idx < len(values) and values[idx] is not None:
                cell = values[idx]
                if hasattr(cell, "isoformat"):
                    row[col] = cell.isoformat()
                else:
                    row[col] = str(cell)
            else:
                row[col] = ""
        rows.append(row)
    return rows, fieldnames


# ── shared row processing ────────────────────────────────────────────


def _error_result(message: str) -> ImportResult:
    r = ImportResult()
    r.errors.append(RowError(0, "", message))
    return r


def _import_rows(rows, fieldnames, *, dry_run, user) -> ImportResult:
    result = ImportResult()
    missing = REQUIRED_COLUMNS - set(fieldnames)
    if missing:
        result.errors.append(RowError(0, "", f"Missing required columns: {sorted(missing)}."))
        return result

    result.rows_total = len(rows)
    if not rows:
        return result

    categories = {c.name: c for c in RequirementCategory.objects.all()}
    sources = {s.name: s for s in RequirementSource.objects.all()}
    levels = {lv.code: lv for lv in RequirementLevel.objects.all()}
    features = {f.name: f for f in Feature.objects.all()}
    projects = {p.name: p for p in Project.objects.all()}
    existing_requirements = {r.identifier: r for r in Requirement.objects.all()}

    try:
        from tcms.management.models import Product  # noqa: WPS433
        products = {p.name: p for p in Product.objects.all()}
    except (ImportError, Exception):  # noqa: BLE001
        products = {}

    sid = transaction.savepoint()
    try:
        for row_num, row in enumerate(rows, start=2):
            identifier = (row.get("identifier") or "").strip()
            title = (row.get("title") or "").strip()
            if not identifier:
                result.errors.append(RowError(row_num, "", "Empty identifier."))
                result.rows_skipped += 1
                continue
            if not title:
                result.errors.append(RowError(row_num, identifier, "Empty title."))
                result.rows_skipped += 1
                continue

            try:
                fk_values = _resolve_fks(
                    row,
                    categories=categories,
                    sources=sources,
                    levels=levels,
                    products=products,
                    projects=projects,
                    features=features,
                    existing_requirements=existing_requirements,
                )
            except ValueError as exc:
                result.errors.append(RowError(row_num, identifier, str(exc)))
                result.rows_skipped += 1
                continue

            try:
                defaults = _build_defaults(row, fk_values)
            except ValueError as exc:
                result.errors.append(RowError(row_num, identifier, str(exc)))
                result.rows_skipped += 1
                continue

            if not dry_run and user is not None and "created_by" not in defaults:
                defaults["created_by"] = user

            obj, created = Requirement.objects.update_or_create(
                identifier=identifier,
                defaults=defaults,
            )
            if created:
                result.rows_created += 1
            else:
                result.rows_updated += 1
            result.rows_ok += 1
            existing_requirements[identifier] = obj

            # Test-case links — only sync when the column was emitted.
            # Empty value clears existing links; absent column leaves
            # them alone (lets users edit one column without dropping
            # links from rows they didn't touch).
            if "linked_cases" in fieldnames:
                try:
                    case_ids = _parse_case_ids(row.get("linked_cases"))
                    _sync_case_links(obj, case_ids, user=user)
                except ValueError as exc:
                    result.errors.append(RowError(row_num, identifier, str(exc)))

        if dry_run:
            transaction.savepoint_rollback(sid)
        else:
            transaction.savepoint_commit(sid)
    except Exception:
        transaction.savepoint_rollback(sid)
        raise

    return result


def _resolve_fks(row, *, categories, sources, levels, products, projects, features, existing_requirements) -> dict:
    out = {}
    out["category"] = _pick(row.get("category"), categories, "category")
    out["source"] = _pick(row.get("source"), sources, "source")
    out["level"] = _pick(row.get("level"), levels, "level (expected slug like 'system')")
    out["product"] = _pick(row.get("product"), products, "product") if products else None
    out["project"] = _pick(row.get("project"), projects, "project")
    out["feature"] = _pick(row.get("feature"), features, "feature")
    out["parent_requirement"] = _pick(
        row.get("parent_requirement"), existing_requirements, "parent_requirement (identifier)"
    )
    out["superseded_by"] = _pick(
        row.get("superseded_by"), existing_requirements, "superseded_by (identifier)"
    )
    return out


def _pick(name, cache, label):
    if name is None:
        return None
    name = str(name).strip()
    if not name:
        return None
    obj = cache.get(name)
    if obj is None:
        raise ValueError(f"Unknown {label}: {name!r}")
    return obj


def _parse_case_ids(value) -> list:
    """Parse a comma-separated list of TestCase IDs from a CSV cell.

    Accepts bare ints (``42``), JIRA-style prefixes (``TC-42``), and any
    mix of whitespace + commas. Returns a list of unique ints in input
    order. Raises ValueError for malformed tokens so the row's error
    surfaces in the per-row report.
    """
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    out = []
    seen = set()
    for token in text.split(","):
        token = token.strip()
        if not token:
            continue
        match = _CASE_ID_PATTERN.match(token)
        if not match:
            raise ValueError(
                f"Invalid TestCase reference {token!r} in linked_cases — "
                "expected integer or 'TC-N' form."
            )
        case_id = int(match.group(1))
        if case_id in seen:
            continue
        seen.add(case_id)
        out.append(case_id)
    return out


def _sync_case_links(requirement, case_ids, *, user=None) -> None:
    """Replace ``requirement.case_links`` so it exactly matches case_ids.

    Verifies every requested TestCase id exists before mutating anything
    so a typo doesn't leave the requirement with half its links removed.
    """
    if not case_ids:
        # Empty linked_cases column → drop all links for this requirement.
        requirement.case_links.all().delete()
        return

    from django.apps import apps  # noqa: WPS433 — lazy
    try:
        TestCase = apps.get_model("testcases", "TestCase")
    except LookupError as exc:
        raise ValueError(f"testcases.TestCase model not available: {exc}")

    found_ids = set(
        TestCase.objects.filter(pk__in=case_ids).values_list("pk", flat=True)
    )
    missing = [str(cid) for cid in case_ids if cid not in found_ids]
    if missing:
        raise ValueError(
            f"Unknown TestCase id(s) in linked_cases: {', '.join(missing)}"
        )

    existing = {
        link.case_id: link
        for link in requirement.case_links.all()
    }
    desired = set(case_ids)

    # Delete links that aren't desired anymore.
    to_delete = [link.pk for cid, link in existing.items() if cid not in desired]
    if to_delete:
        RequirementTestCaseLink.objects.filter(pk__in=to_delete).delete()

    # Create links that are desired but don't yet exist.
    to_create = [cid for cid in case_ids if cid not in existing]
    for case_id in to_create:
        kwargs = {
            "requirement": requirement,
            "case_id": case_id,
            "link_type": "verifies",
            "suspect": False,
        }
        if user is not None and user.is_authenticated:
            kwargs["created_by"] = user
        RequirementTestCaseLink.objects.create(**kwargs)


def _build_defaults(row, fk_values) -> dict:
    def cell(name, default=""):
        value = row.get(name)
        if isinstance(value, str):
            return value.strip()
        return value if value is not None else default

    defaults = {
        "title": cell("title"),
        "description": cell("description"),
        "rationale": cell("rationale"),
        "source_section": cell("source_section"),
        "document_file_name": cell("document_file_name"),
        "document_title": cell("document_title"),
        "status": cell("status") or "draft",
        "priority": cell("priority") or "medium",
        "verification_method": cell("verification_method") or "test",
        "verification_exemption_reason": cell("verification_exemption_reason"),
        "asil": cell("asil"),
        "sil": cell("sil"),
        "iec62304_class": cell("iec62304_class"),
        "dal": cell("dal"),
        "doc_id": cell("doc_id"),
        "doc_revision": cell("doc_revision"),
        "change_reason": cell("change_reason"),
        "jira_issue_key": cell("jira_issue_key"),
    }
    effective_date = cell("effective_date")
    if effective_date:
        defaults["effective_date"] = effective_date  # Django parses ISO date strings
    external_refs = cell("external_refs")
    if external_refs:
        try:
            defaults["external_refs"] = json.loads(external_refs)
        except json.JSONDecodeError:
            raise ValueError(f"external_refs must be valid JSON, got: {external_refs!r}")
    defaults.update({k: v for k, v in fk_values.items() if v is not None})
    return defaults
