"""v0.4.x: document_file_name + document_title round-trip through exports + form.

Run via:

    manage.py test tcms_requirements.tests.test_document_fields --settings=tcms.settings.test
"""
import io

from django.contrib.auth import get_user_model
from django.test import TestCase

from tcms.management.models import Classification, Product

from tcms_requirements.exports.csv_export import COLUMNS, write_csv
from tcms_requirements.exports.json_export import _requirement_payload
from tcms_requirements.exports.templates import (
    SAMPLE_ROWS,
    build_xlsx_template,
    write_csv_template,
)
from tcms_requirements.forms import RequirementForm
from tcms_requirements.imports.csv_import import import_bytes
from tcms_requirements.models import (
    Project,
    Requirement,
    RequirementLevel,
)


class _DocFieldsSetup(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.classification = Classification.objects.create(name="DocFields")
        cls.product = Product.objects.create(
            name="Doc Fields Product", classification=cls.classification,
        )
        cls.user = get_user_model().objects.create_user(
            username="docs-tester", email="docs@example.com",
        )
        cls.level = RequirementLevel.objects.create(
            code="SYS", name="System", order=1,
        )
        cls.project = Project.objects.create(
            product=cls.product, name="Doc Project", code="DP",
        )
        cls.requirement = Requirement.objects.create(
            product=cls.product, project=cls.project, level=cls.level,
            identifier="DOC-1", title="Document fields demo",
            status="approved",
            document_file_name="system-spec_v2.pdf",
            document_title="System Requirements Specification — Platform 2026",
            created_by=cls.user,
        )


class CSVColumnsTest(_DocFieldsSetup):
    def test_csv_columns_include_document_fields(self):
        self.assertIn("document_file_name", COLUMNS)
        self.assertIn("document_title", COLUMNS)
        # Both appear before `level` in the taxonomy block.
        self.assertLess(COLUMNS.index("document_file_name"), COLUMNS.index("level"))
        self.assertLess(COLUMNS.index("document_title"), COLUMNS.index("level"))


class CSVExportRoundTripTest(_DocFieldsSetup):
    def test_csv_export_contains_document_field_values(self):
        buf = io.StringIO()
        write_csv([self.requirement], buf)
        output = buf.getvalue()
        self.assertIn("system-spec_v2.pdf", output)
        self.assertIn("System Requirements Specification — Platform 2026", output)


class JSONExportRoundTripTest(_DocFieldsSetup):
    def test_json_payload_contains_document_fields(self):
        payload = _requirement_payload(self.requirement)
        self.assertEqual(payload["document_file_name"], "system-spec_v2.pdf")
        self.assertEqual(
            payload["document_title"],
            "System Requirements Specification — Platform 2026",
        )


class ImportTemplateTest(_DocFieldsSetup):
    def test_csv_template_emits_document_field_columns(self):
        buf = io.StringIO()
        write_csv_template(buf)
        text = buf.getvalue()
        # Header row + at least one sample row populates the new columns.
        self.assertIn("document_file_name", text)
        self.assertIn("document_title", text)
        self.assertIn("customer-rfp_v2.pdf", text)


class XLSXTemplateTest(_DocFieldsSetup):
    def test_xlsx_template_includes_document_fields_in_header_and_samples(self):
        from openpyxl import load_workbook  # noqa: WPS433 — test-local

        payload = build_xlsx_template()
        wb = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[0])

        self.assertIn("document_file_name", header)
        self.assertIn("document_title", header)

        file_idx = header.index("document_file_name")
        title_idx = header.index("document_title")
        sample_file_values = {row[file_idx] for row in rows[1:] if row[file_idx]}
        sample_title_values = {row[title_idx] for row in rows[1:] if row[title_idx]}
        self.assertIn("customer-rfp_v2.pdf", sample_file_values)
        self.assertIn("Customer RFP — Platform 2026", sample_title_values)


class XLSXImportRoundTripTest(_DocFieldsSetup):
    def test_xlsx_import_persists_document_fields(self):
        from openpyxl import Workbook  # noqa: WPS433 — test-local

        wb = Workbook()
        ws = wb.active
        ws.append(["identifier", "title", "document_file_name", "document_title"])
        ws.append([
            "DOC-XLSX-1",
            "XLSX import demo",
            "imported-spec.xlsx",
            "Imported Specification — Round-trip",
        ])
        buf = io.BytesIO()
        wb.save(buf)

        result = import_bytes(buf.getvalue(), "fixtures.xlsx", dry_run=False, user=self.user)

        self.assertEqual(result.errors, [])
        self.assertEqual(result.rows_created, 1)
        instance = Requirement.objects.get(identifier="DOC-XLSX-1")
        self.assertEqual(instance.document_file_name, "imported-spec.xlsx")
        self.assertEqual(instance.document_title, "Imported Specification — Round-trip")


class FormPersistsDocumentFieldsTest(_DocFieldsSetup):
    def test_form_persists_document_fields(self):
        form = RequirementForm(data={
            "identifier": "DOC-2",
            "title": "Form-driven requirement",
            "description": "",
            "rationale": "",
            "category": "",
            "source": "",
            "source_section": "",
            "document_file_name": "form-spec.pdf",
            "document_title": "Form-driven Document",
            "level": self.level.pk,
            "product": self.product.pk,
            "project": self.project.pk,
            "feature": "",
            "parent_requirement": "",
            "status": "draft",
            "priority": "medium",
            "verification_method": "test",
            "verification_exemption_reason": "",
            "asil": "",
            "sil": "",
            "iec62304_class": "",
            "dal": "",
            "doc_id": "",
            "doc_revision": "",
            "effective_date": "",
            "superseded_by": "",
            "change_reason": "",
            "jira_issue_key": "",
        })

        self.assertTrue(form.is_valid(), msg=form.errors)
        instance = form.save(commit=False)
        instance.created_by = self.user
        instance.save()

        instance.refresh_from_db()
        self.assertEqual(instance.document_file_name, "form-spec.pdf")
        self.assertEqual(instance.document_title, "Form-driven Document")
